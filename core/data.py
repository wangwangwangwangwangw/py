'''
#!/usr/bin/python3
# author: Hwei
# 2023年05月07日
# Email: xxx@qq.com
用于加载文件
'''
import csv
from openpyxl import load_workbook
from pathlib import Path
#从当前文件（__file__），的上一级目录的上一级目录（parent.parent）的data包目录
data_path=Path(__file__).parent.parent /"data"

#因为代码要允许不同用例读取不同文件所以通过参数方式（filename）支持多个方式
def read_csv(filename):
    path =data_path / filename

    with open(path,encoding='utf-8') as f:
        reader = csv.DictReader(f)#DictReader  是csv的文件读取器
        for i in reader:
            yield list(i.values())#重点,便读取出来的值变成列表形式 yield把值返回给了调用者

def read_excel(filename):
    path = data_path / filename
    wb =load_workbook(path)#加载数据表的方法只需呀传入路径
    ws =wb.active#选择默认sheet页
    yield from ws.iter_rows(min_row=2,values_only=True)#迭代器；第一行是表头 从第二行开始（min_row=2）,valuers_only= True是只要数据表里面的值不要颜色字体等等

def read_esshop(filename):
    path = data_path / filename
    wb = load_workbook(path)
    #获取指定名称的折页：
    wb = wb['Sheet3']
    # 迭代器；第一行是表头 从第二行开始（min_row=2）,valuers_only= True是只要数据表里面的值不要颜色字体等等
    yield from wb.iter_rows(min_row=2,values_only=True)




if __name__ == '__main__':
    '''重点,便读取出来的值变成列表形式打印
    ['standard_user', 'secret_sauce']
    ['standard', 'secret_sauce']
    '''
    for d in read_esshop('ddt_xinzenghuiyuan.xlsx'):
        print(d)